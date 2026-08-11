import io
import os
import re
import sys
import threading
import uuid
from collections import OrderedDict

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from flask import Flask, jsonify, render_template, request, send_file

from config import CACHE
from core import brain as brain_mod
from core import cell
from core import route
from core import weight
from data import loader
from data import warehouse
from tools import heatmap
from web import dashboard
from web import router

MAX_JOBS = 20
CAP_UNLIMITED = 10**9
MAX_SUGGEST = 10
MAX_BATCH = 200

app = Flask(__name__)
app.config["TEMPLATES_AUTO_RELOAD"] = True
app.config["MAX_CONTENT_LENGTH"] = 2 * 1024**3

JOBS = OrderedDict()
JOBS_LOCK = threading.Lock()


class Job:
    def __init__(self):
        self.id = uuid.uuid4().hex[:12]
        self.progress = []
        self.status = "running"
        self.result = None
        self.error = None

    def log(self, message):
        self.progress.append(message)


def register(job):
    with JOBS_LOCK:
        JOBS[job.id] = job
        while len(JOBS) > MAX_JOBS:
            JOBS.popitem(last=False)


def get_job(job_id):
    with JOBS_LOCK:
        return JOBS.get(job_id)


def run_job(job, orders, frequency, current_aisles, aisle_capacity, name, category, alias=None):
    try:
        brain = brain_mod.build(
            orders, frequency, current_aisles, aisle_capacity, name, category, log=job.log
        )
        brain["alias"] = alias or {}
        job.log("Saving model...")
        brain_mod.save(brain)
        job.result = {
            "advice": brain["advice"],
            "gu_map": brain["gu_map"],
            "stats": brain["stats"],
            "meta": brain["meta"],
        }
        job.status = "done"
        job.log("Done. Cached for faster lookups.")
    except Exception as exc:
        job.error = str(exc)
        job.status = "error"
        job.log(f"Error: {exc}")


@app.route("/")
def index():
    return render_template("index.html")


@app.after_request
def _no_cache(resp):
    resp.headers["Cache-Control"] = "no-store, must-revalidate"
    resp.headers["Pragma"] = "no-cache"
    return resp


_SEARCHABLE = None
_SKU_NUMBERS = None


def _sku_numbers(brain):
    global _SKU_NUMBERS
    if _SKU_NUMBERS is not None:
        return _SKU_NUMBERS
    shortest = {}
    for code, internal in brain.get("alias", {}).items():
        if not code.isdigit():
            continue
        current = shortest.get(internal)
        if current is None or (len(code), code) < (len(current), current):
            shortest[internal] = code
    _SKU_NUMBERS = shortest
    return _SKU_NUMBERS


def _searchable_codes(brain):
    global _SEARCHABLE
    if _SEARCHABLE is not None:
        return _SEARCHABLE
    advice = brain["advice"]
    out, seen = [], set()
    for code, internal in brain.get("alias", {}).items():
        if internal not in advice:
            continue
        name = advice[internal].get("name", "")
        shown = "SKU" + code if code.isdigit() else code
        if (internal, shown) in seen:
            continue
        seen.add((internal, shown))
        out.append((code, shown, name, internal))
    for internal, adv in advice.items():
        name = adv.get("name") or ""
        if name and (internal, name) not in seen:
            seen.add((internal, name))
            out.append((name, name, name, internal))
    _SEARCHABLE = out
    return _SEARCHABLE


def _strip_sku_prefix(text):
    return re.sub(r"(?i)^sku[-\s]?", "", text)


def _autocomplete(query, rows):
    query = query.upper()
    head, rest, seen = [], [], set()
    for code, shown, name, internal in rows:
        if query not in code.upper():
            continue
        if internal in seen:
            continue
        seen.add(internal)
        item = {"code": shown, "name": name, "id": internal}
        (head if code.upper().startswith(query) else rest).append(item)
        if len(head) >= MAX_SUGGEST:
            break
    return (head + rest)[:MAX_SUGGEST]


@app.route("/api/suggest")
def suggest():
    query = _strip_sku_prefix((request.args.get("q") or "").strip())
    if len(query) < 2:
        return jsonify([])
    brain = brain_mod.load()
    if brain is None:
        return jsonify([])
    return jsonify(_autocomplete(query, _searchable_codes(brain)))


@app.route("/api/refresh-stock", methods=["POST"])
def refresh_stock():
    try:
        cell.load(force=True)
        return jsonify({"ok": True})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/reset-session", methods=["POST"])
def reset_session():
    try:
        cell.reset_session()
        return jsonify({"ok": True})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/all-skus")
def all_skus():
    brain = brain_mod.load()
    if brain is None:
        return jsonify({"items": [], "cats": []})
    internal_to_num = {}
    for code, internal in brain.get("alias", {}).items():
        if code.isdigit() and internal not in internal_to_num:
            internal_to_num[internal] = code
    items, cats = [], set()
    for internal, a in brain["advice"].items():
        num = internal_to_num.get(internal)
        if not num:
            continue
        cat = a.get("category") or "Unknown category"
        cats.add(cat)
        items.append(
            {
                "code": "SKU" + num,
                "name": a.get("name", ""),
                "id": internal,
                "cat": cat,
                "pick": a.get("pick_count", 0),
            }
        )
    items.sort(key=lambda x: -x["pick"])
    return jsonify({"items": items, "cats": sorted(cats)})


_LAYOUT = None


@app.route("/api/layout")
def layout():
    if _not_ready():
        return jsonify({})
    global _LAYOUT
    if _LAYOUT is None:
        data = cell.load()
        zones = {}
        for aisle in data:
            m = re.match(r"^([A-Za-z]+)(\d+)$", aisle)
            if not m:
                continue
            zone, aisle_num = m.group(1), int(m.group(2))
            zones.setdefault(zone, []).append((aisle_num, aisle))
        _LAYOUT = {z: [a for _, a in sorted(v)] for z, v in zones.items()}
    return jsonify(_LAYOUT)


@app.route("/api/brain")
def brain_status():
    brain = brain_mod.load()
    if brain is None:
        return jsonify({"ready": False})
    return jsonify(
        {
            "ready": True,
            "meta": brain.get("meta", {}),
            "stats": brain.get("stats", {}),
            "gu_map": brain.get("gu_map", []),
        }
    )


@app.route("/api/place", methods=["POST"])
def place():
    brain = brain_mod.load()
    if brain is None:
        return jsonify({"error": "No model yet. Load order history first."}), 400

    data = request.get_json(force=True)
    items = [it for it in data.get("items", []) if it.get("item")]
    if not items:
        return jsonify({"error": "No item codes given."}), 400

    _seen = set()
    _unique = []
    for it in items:
        key = brain_mod._resolve(brain, it["item"]) or it["item"]
        if key in _seen:
            continue
        _seen.add(key)
        _unique.append(it)
    items = _unique
    extra = max(0, len(items) - MAX_BATCH)
    items = items[:MAX_BATCH]

    def pick_of(it):
        resolved = brain_mod._resolve(brain, it["item"])
        return brain["advice"].get(resolved, {}).get("pick_count", 0)

    order = sorted(range(len(items)), key=lambda i: -pick_of(items[i]))

    cell.release({brain_mod._resolve(brain, it["item"]) or it["item"] for it in items})
    used_cells = set()
    group_aisles = set()
    placed = {}
    for i in order:
        it = items[i]

        comps = it.get("companions") or []
        comp_ids = [c.get("name") or c.get("id") if isinstance(c, dict) else c for c in comps]
        placed[i] = brain_mod.place(
            brain,
            it["item"],
            companion_ids=[c for c in comp_ids if c],
            category=it.get("category"),
            name=it.get("name"),
            used_cells=used_cells,
            group_aisles=group_aisles,
        )
    results = [placed[i] for i in range(len(items))]

    brain_mod.tune_batch(results, brain=brain, used_cells=used_cells)
    sku_numbers = _sku_numbers(brain)
    for r in results:
        r.pop("_reserve_days", None)
        internal = r.pop("_internal", None)
        if internal is not None:
            r["id"] = internal
            number = sku_numbers.get(internal)
            if number:
                r["sku"] = "SKU" + number
    tally = _tally_confidence(results)
    resp = {"results": results, "tally": tally, "meta": brain.get("meta", {})}
    if extra:
        resp["warning"] = (
            f"Batch too large. Only the first {MAX_BATCH} codes were processed ({extra} skipped)."
        )
    return jsonify(resp)


def _tally_confidence(results):
    tally = {}
    for r in results:
        tally[r["confidence"]] = tally.get(r["confidence"], 0) + 1
    return tally


NOT_READY_MSG = ""


def _not_ready():
    ok, _ = warehouse.is_ready()
    return not ok


@app.route("/api/load-history", methods=["POST"])
def load_history():
    files = list(request.files.values())
    if not files:
        return jsonify({"error": "No files selected."}), 400

    accepted = {}
    skipped = []
    for f in files:
        raw = f.read()
        role = loader.detect_role(raw)
        if role and role not in accepted:
            f.stream.seek(0)
            accepted[role] = (f, raw)
        else:
            skipped.append(f.filename or "?")

    missing = [v for v in warehouse.ROLE_FILES if v not in accepted]
    if missing:
        labels = {
            "pick": "order history",
            "refill_log": "refill log",
            "location": "locations",
            "stock": "stock",
            "product": "products",
        }
        return (
            jsonify(
                {
                    "error": "Missing files: "
                    + ", ".join(labels[v] for v in missing)
                    + ". Drop in all 5 types."
                }
            ),
            400,
        )

    try:
        warehouse.WAREHOUSE_DIR.mkdir(parents=True, exist_ok=True)
        for role, (f, raw) in accepted.items():
            (warehouse.WAREHOUSE_DIR / warehouse.ROLE_FILES[role]).write_bytes(raw)
        warehouse.write_config(dict(warehouse.DEFAULT_COLS), dict(warehouse.DEFAULT_VALS))
    except OSError as exc:
        return jsonify({"error": f"Could not save warehouse files: {exc}"}), 500

    for f in CACHE.glob("*.json"):
        f.unlink()
    if brain_mod.exists():
        brain_mod.BRAIN_FILE.unlink()
    _reset_runtime_caches()

    detected = {v: accepted[v][0].filename for v in accepted}
    named = [(accepted[v][0].filename or v, accepted[v][1]) for v in ("pick", "location", "product")]
    diag = router.diagnose(named)
    job = Job()
    register(job)
    job.log("5 files identified and saved. Learning slotting from order history...")
    try:
        orders, frequency, current_aisles = loader.load_pick_history(
            diag["pick_bytes"], diag["location_aisle"], log=job.log
        )
    except loader.LoadError as exc:
        job.status = "error"
        job.error = str(exc)
        return jsonify({"ok": True, "job": job.id, "detected": detected, "skipped": skipped})
    aisle_capacity = diag["aisle_capacity"] or {"A1": CAP_UNLIMITED}
    threading.Thread(
        target=run_job,
        args=(
            job,
            orders,
            frequency,
            current_aisles,
            aisle_capacity,
            diag["name"],
            diag["category"],
            diag["alias"],
        ),
        daemon=True,
    ).start()
    return jsonify({"ok": True, "job": job.id, "detected": detected, "skipped": skipped})


@app.route("/api/clear", methods=["POST"])
def clear_data():
    import shutil

    if warehouse.WAREHOUSE_DIR.exists():
        shutil.rmtree(str(warehouse.WAREHOUSE_DIR), ignore_errors=True)
    for f in list(CACHE.glob("*.json")) + list(CACHE.glob("*.pkl")):
        f.unlink()
    warehouse._config = None
    _reset_runtime_caches()
    return jsonify({"ok": True})


def _reset_runtime_caches():
    global _LAYOUT, _SEARCHABLE, _SKU_NUMBERS
    cell.reset_session()
    weight._WEIGHTS = None
    weight._SORTED = None
    _LAYOUT = None
    _SEARCHABLE = None
    _SKU_NUMBERS = None


@app.route("/api/dashboard")
def dashboard_data():
    if _not_ready():
        return jsonify({"error": NOT_READY_MSG, "not_ready": True}), 200
    try:
        return jsonify(dashboard.load())
    except Exception as exc:
        return jsonify({"error": f"Could not read dashboard data: {exc}"}), 500


@app.route("/api/heat")
def heat():
    """Activity per aisle: picks, refills out, refills in."""
    if _not_ready():
        return jsonify({"error": NOT_READY_MSG, "not_ready": True}), 200
    try:
        return jsonify(heatmap.load())
    except Exception as exc:
        return jsonify({"error": f"Could not read the heatmap: {exc}"}), 500


@app.route("/api/route", methods=["POST"])
def route_of_order():
    """Measure travel for one order, before and after the location change.

    The warehouse does not route pickers; this is a measurement for comparison.
    """
    data = request.get_json(force=True)
    results = [r for r in data.get("results", []) if r.get("aisle")]
    if not results:
        return jsonify({"error": "Nothing to measure yet."}), 400

    brain = brain_mod.load()
    after = [r["aisle"] for r in results]
    before = []
    for r in results:
        advice = (brain or {}).get("advice", {}).get(r.get("item"), {})
        before.append(advice.get("current_aisle") or r.get("pick_aisle") or r["aisle"])

    return jsonify({
        "before": route.measure_order(before),
        "after": route.measure_order(after),
        "note": "Measured in aisles. This measures distance, it does not route pickers.",
    })


@app.route("/api/status/<job_id>")
def status(job_id):
    job = get_job(job_id)
    if job is None:
        return jsonify({"error": "Job not found."}), 404
    body = {"status": job.status, "progress": job.progress}
    if job.status == "done":
        body["result"] = job.result
    if job.status == "error":
        body["error"] = job.error
    return jsonify(body)


def _floor_number(floor):
    m = re.search(r"\d+", floor or "")
    return m.group(0) if m else ""


# Column titles the target WMS expects in its import sheet. They stay in the
# original language because the importer matches on the exact header text.
EXCEL_HEADER = (
    "Mã Dãy Kệ",
    "Dãy kệ",
    "Tầng",
    "Mã Sản Phẩm",
    "Tên Sản Phẩm",
    "Vị trí",
    "Seq Min",
    "Seq Max",
    "Ghi Chú",
)
EXCEL_WIDTHS = (14, 12, 8, 16, 22, 12, 10, 10, 30)


def _xlsx_from_rows(rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(list(EXCEL_HEADER))
    for header_cell in ws[1]:
        header_cell.font = Font(bold=True)
    for idx, width in enumerate(EXCEL_WIDTHS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width

    for aisle, item, name, floor, cell_code in rows:
        ws.append(
            ["", aisle, _floor_number(floor), item, name or "", cell_code or "", "", "", floor or ""]
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.route("/api/excel-place", methods=["POST"])
def excel_place():
    data = request.get_json(force=True)
    results = [r for r in data.get("results", []) if r.get("aisle")]
    if not results:
        return jsonify({"error": "Nothing to export yet."}), 400
    rows = sorted(
        (
            (r["aisle"], r.get("item"), r.get("name"), r.get("floor"), r.get("cell"))
            for r in results
        ),
        key=lambda row: row[0],
    )
    return send_file(
        _xlsx_from_rows(rows),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="suggested_locations.xlsx",
    )


if __name__ == "__main__":
    debug = os.environ.get("SLOTWISE_DEBUG", "") == "1"
    app.run(
        host=os.environ.get("SLOTWISE_HOST", "127.0.0.1"),
        port=int(os.environ.get("SLOTWISE_PORT", "5000")),
        debug=debug,
        use_reloader=debug,
        threaded=True,
    )